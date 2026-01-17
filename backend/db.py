import os
from typing import Any, Dict, List, Optional
from openpyxl import Workbook, load_workbook

DB_PATH = os.getenv("EXCEL_DB_PATH", os.path.join(os.getcwd(), "sales_orders.xlsx"))

HEADER_SHEET = "SalesOrderHeader"
DETAIL_SHEET = "SalesOrderDetail"

HEADER_COLUMNS = [
    "SalesOrderNumber",
    "CustomerName",
    "OrderDate",
    "DueDate",
    "BillTo",
    "ShipTo",
    "SubTotal",
    "TaxAmt",
    "Freight",
    "TotalDue",
]

DETAIL_COLUMNS = [
    "SalesOrderNumber",
    "LineNumber",
    "Product",
    "OrderQty",
    "UnitPrice",
    "LineTotal",
]


def _ensure_workbook():
    if os.path.exists(DB_PATH):
        return

    wb = Workbook()
    ws_h = wb.active
    ws_h.title = HEADER_SHEET
    ws_h.append(HEADER_COLUMNS)

    ws_d = wb.create_sheet(DETAIL_SHEET)
    ws_d.append(DETAIL_COLUMNS)

    wb.save(DB_PATH)


def init_db():
    _ensure_workbook()


def _safe_num(v):
    try:
        if v is None or v == "":
            return None
        return float(v)
    except Exception:
        return None


def _safe_int(v):
    try:
        if v is None or v == "":
            return None
        return int(v)
    except Exception:
        return None


def save_order(payload: Dict[str, Any]) -> None:
    """
    Upserts by SalesOrderNumber:
      - Header row: replace if exists, else append
      - Detail rows: delete existing for that order, then append new rows
    """
    _ensure_workbook()
    wb = load_workbook(DB_PATH)
    ws_h = wb[HEADER_SHEET]
    ws_d = wb[DETAIL_SHEET]

    header = payload.get("SalesOrderHeader") or {}
    details = payload.get("SalesOrderDetail") or []

    sales_order_number = header.get("SalesOrderNumber")
    if not sales_order_number:
        raise ValueError("SalesOrderHeader.SalesOrderNumber is required")

    # --- Upsert header ---
    existing_row_idx = None
    for r in range(2, ws_h.max_row + 1):
        if ws_h.cell(row=r, column=1).value == sales_order_number:
            existing_row_idx = r
            break

    row_values = [
        sales_order_number,
        header.get("CustomerName"),
        header.get("OrderDate"),
        header.get("DueDate"),
        header.get("BillTo"),
        header.get("ShipTo"),
        _safe_num(header.get("SubTotal")),
        _safe_num(header.get("TaxAmt")),
        _safe_num(header.get("Freight")),
        _safe_num(header.get("TotalDue")),
    ]

    if existing_row_idx:
        for c, v in enumerate(row_values, start=1):
            ws_h.cell(row=existing_row_idx, column=c).value = v
    else:
        ws_h.append(row_values)

    # --- Replace details for this order ---
    # delete rows from bottom to top to avoid index shifting
    rows_to_delete = []
    for r in range(2, ws_d.max_row + 1):
        if ws_d.cell(row=r, column=1).value == sales_order_number:
            rows_to_delete.append(r)
    for r in reversed(rows_to_delete):
        ws_d.delete_rows(r, 1)

    for idx, it in enumerate(details, start=1):
        ws_d.append([
            sales_order_number,
            idx,
            it.get("Product"),
            _safe_int(it.get("OrderQty")),
            _safe_num(it.get("UnitPrice")),
            _safe_num(it.get("LineTotal")),
        ])

    wb.save(DB_PATH)


def list_orders() -> List[Dict[str, Any]]:
    _ensure_workbook()
    wb = load_workbook(DB_PATH)
    ws_h = wb[HEADER_SHEET]

    out = []
    for r in range(2, ws_h.max_row + 1):
        row = {HEADER_COLUMNS[c - 1]: ws_h.cell(row=r, column=c).value for c in range(1, len(HEADER_COLUMNS) + 1)}
        if row.get("SalesOrderNumber"):
            out.append(row)
    return out


def get_order(sales_order_number: str) -> Optional[Dict[str, Any]]:
    _ensure_workbook()
    wb = load_workbook(DB_PATH)
    ws_h = wb[HEADER_SHEET]
    ws_d = wb[DETAIL_SHEET]

    header = None
    for r in range(2, ws_h.max_row + 1):
        if ws_h.cell(row=r, column=1).value == sales_order_number:
            header = {HEADER_COLUMNS[c - 1]: ws_h.cell(row=r, column=c).value for c in range(1, len(HEADER_COLUMNS) + 1)}
            break
    if not header:
        return None

    details = []
    for r in range(2, ws_d.max_row + 1):
        if ws_d.cell(row=r, column=1).value == sales_order_number:
            details.append({
                "Product": ws_d.cell(row=r, column=3).value,
                "OrderQty": ws_d.cell(row=r, column=4).value,
                "UnitPrice": ws_d.cell(row=r, column=5).value,
                "LineTotal": ws_d.cell(row=r, column=6).value,
            })

    return {"SalesOrderHeader": header, "SalesOrderDetail": details}
